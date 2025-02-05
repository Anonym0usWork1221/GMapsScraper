from openpyxl import Workbook, load_workbook
from threading import Lock
from csv import DictWriter
import json
import os


class CSVCreator:
    def __init__(self, file_lock: Lock, output_path: str = "./CSV_FILES"):
        self._output_path = output_path
        self._file_lock = file_lock

    def create(self, list_of_dict_data: list[dict]):
        with self._file_lock:
            os.makedirs(self._output_path, exist_ok=True)
            file_name = "google_maps_data.csv"
            _isheader_file = False
            if not os.path.isfile(self._output_path + "/" + file_name):
                _isheader_file = True

            if _isheader_file:
                file_handler = open(self._output_path + "/" + file_name, "w", newline="", encoding="utf-8-sig")
            else:
                file_handler = open(self._output_path + "/" + file_name, "a", newline="", encoding="utf-8-sig")

            writer = DictWriter(file_handler, fieldnames=list_of_dict_data[0].keys(), extrasaction='ignore')
            if _isheader_file:
                writer.writeheader()

            writer.writerows(list_of_dict_data)
            file_handler.close()


class JSONCreator:
    def __init__(self, file_lock: Lock, output_path: str = "./JSON_FILES"):
        self._file_lock = file_lock
        self._output_path = output_path

    def create(self, list_of_dict_data: list[dict]):
        file_name = "google_maps_data.json"
        file_path = os.path.join(self._output_path, file_name)

        with self._file_lock:
            os.makedirs(self._output_path, exist_ok=True)
            if not os.path.isfile(file_path):
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(list_of_dict_data, f, ensure_ascii=False, indent=4)
            else:
                with open(file_path, 'r+', encoding='utf-8') as f:
                    try:
                        existing_data = json.load(f)
                    except json.JSONDecodeError:
                        existing_data = []
                    existing_data.extend(list_of_dict_data)
                    f.seek(0)
                    f.truncate()
                    json.dump(existing_data, f, ensure_ascii=False, indent=4)



class XLSXCreator:
    def __init__(self, file_lock: Lock, output_path: str = "./XLSX_FILES"):
        self._file_lock = file_lock
        self._output_path = output_path

    def create(self, list_of_dict_data: list[dict]):
        file_name = "google_maps_data.xlsx"
        file_path = os.path.join(self._output_path, file_name)

        with self._file_lock:
            os.makedirs(self._output_path, exist_ok=True)
            if not os.path.isfile(file_path):
                wb = Workbook()
                ws = wb.active
                ws.title = "Data"
                headers = list(list_of_dict_data[0].keys())
                ws.append(headers)
                for data_dict in list_of_dict_data:
                    row_values = [data_dict.get(h, "") for h in headers]
                    ws.append(row_values)

                wb.save(file_path)
            else:
                wb = load_workbook(file_path)
                ws = wb.active
                headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                for data_dict in list_of_dict_data:
                    row_values = [data_dict.get(h, "") for h in headers]
                    ws.append(row_values)
                wb.save(file_path)


if __name__ == '__main__':
    data = [
        {"Name": "Store A", "Address": "123 Main St", "Rating": 4.5},
        {"Name": "Store B", "Address": "456 Elm St", "Rating": 3.7}
    ]

    file_thread_lock = Lock()
    csv_creator = CSVCreator(file_thread_lock, output_path="./CSV_FILES")
    csv_creator.create(data)

    # JSON usage
    json_creator = JSONCreator(file_thread_lock, output_path="./JSON_FILES")
    json_creator.create(data)

    # XLSX usage
    xlsx_creator = XLSXCreator(file_thread_lock, output_path="./XLSX_FILES")
    xlsx_creator.create(data)


